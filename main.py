import openpyxl
import datetime
import telebot
from telebot import types

bot = telebot.TeleBot("5121978271:AAFNK4Pz9Of5-u6Bv0OZfXy2eQktjjs-aag")


def create_shedule():
    for column in range(5,  excel_file['Лист1'].max_column+1):
        lessons = {}
        class_ = excel_file['Лист1'].cell(row=1, column=column).value
        for row in range(2, 12):
            lesson = excel_file['Лист1'].cell(row=row, column=5).value
            time = excel_file['Лист1'].cell(row=row, column=1).value
            try:
                start_hours = int(time[:2])
                start_minutes = int(time[3:-8])
                end_hours = int(time[8:-3])
                end_minutes = int(time[11:])
            except TypeError:
                None
            lessons.update({lesson: {'start': [start_hours, start_minutes], 'end': [end_hours, end_minutes]}})
        shedule.update({class_: lessons})
    return shedule


def get_class_lesson(shedule, class_, current_time):
    for lesson in shedule[class_]:
        start = shedule[class_][lesson]['start']
        end = shedule[class_][lesson]['end']
        lesson_start = datetime.time(start[0], start[1], 0)
        lesson_end = datetime.time(end[0], end[1], 0)
        if time_in_range(lesson_start, lesson_end, current_time):
            return lesson
        elif lesson is None:
            return f"У класса {class_} нет сейчас уроков"


def time_in_range(start, end, current):
    return start <= current <= end


@bot.message_handler(content_types=['text'])
def send_class_lesson(message):
    if message.text == "/start":
        classes = types.ReplyKeyboardMarkup(resize_keyboard=True)
        for class_ in shedule:
            classes.add(class_) #Мог сделать через callback, но кода было бы в разы больше. 
        bot.send_message(message.from_user.id, "Выберите класс: ", reply_markup=classes)
    else:
        try:
            lesson = get_class_lesson(shedule, message.text, current)
            bot.send_message(message.from_user.id, f"У класса {message.text} сейчас {lesson}")
        except KeyError:
            bot.send_message(message.from_user.id, "Такого класса нету в расписании")

if __name__ == "__main__":
    excel_file = openpyxl.load_workbook('shedule.xlsx')
    shedule = {}
    current = datetime.datetime.now().time()
    shedule = create_shedule()

    bot.polling()
